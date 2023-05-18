from statistics import mean

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from date_mark.not_db_models import GradeModel, SubjectModel, StudentModel, DateModel, StudentProblemsModel, \
    ProblemModel


class ImportService:
    import_file_path = 'date_mark/input/'
    format_input = '.xlsx'

    def import_data(self, filename) -> GradeModel:
        path = self.import_file_path + filename + self.format_input

        # Load in the workbooks
        wb = load_workbook(path)

        # Get a sheet by name
        sheet = wb.get_sheet_by_name(filename)
        classname = sheet['A7'].value.lower().split(': ')[1]
        grade_model = GradeModel(name=classname, subjects=[])
        year = sheet['A5'].value.split('Учебный год: ')[1].lower().split('/')[0]

        subject_model = []

        for i in range(2, 2000):
            column_index = column_index_from_string('A')
            student_number = str(sheet[get_column_letter(column_index) + str(i)].value)

            if 'Предмет:' in student_number:
                subject_name = student_number.split('Предмет:')[1].lower().split('/' + classname)
                subject_model = SubjectModel(name=subject_name[0] + subject_name[1])

            student_count = 0
            date_count = 3
            if student_number == "№":
                j = i + 2
                while sheet["A" + str(j)].value:
                    j += 1
                student_count = j - i  # 1 лишний раз, а 2 прибавленные в начале учтены позже

                while sheet[get_column_letter(date_count) + str(i + 1)].value:
                    date_count += 1

                for j in range(i + 2, i + student_count):
                    student_model = StudentModel(id=sheet['A' + str(j)].value, name=sheet['B' + str(j)].value)

                    month = ''
                    for d in range(3, date_count):
                        degree = sheet[get_column_letter(d) + str(j)].value
                        day = sheet[get_column_letter(d) + str(i + 1)].value
                        month_may = sheet[get_column_letter(d) + str(i)].value
                        if month_may:
                            month = month_may
                        if day.isdigit():
                            degrees = [degree]
                            if not self.check_int(degree) and degree is not None and degree != 'Н':
                                degrees = degree.split()

                            for deg in degrees:
                                type_of_work_temp = sheet[get_column_letter(5) + str(i + student_count + 2 + d)].value
                                date = sheet[get_column_letter(1) + str(i + student_count + 2 + d)].value
                                theme = sheet[get_column_letter(2) + str(i + student_count + 2 + d)].value
                                d += 1
                                date_model = DateModel(degree=deg, day=day, month=month, theme=theme,
                                                       type_of_work=type_of_work_temp, date=date)
                                date_model.year = year
                                student_model.dates.append(date_model)
                        else:
                            date_model = DateModel(degree=degree, day=day, month=month, theme='Итоговая',
                                                   type_of_work='Итоговая', date=day)
                            date_model.year = year
                            student_model.dates.append(date_model)
                    subject_model.students.append(student_model)
                grade_model.subjects.append(subject_model)
        return grade_model

    def get_recommend(self, filename: str) -> str:
        grade_model = self.import_data(filename)
        response = '<h1>-------------------' + grade_model.name + '-------------------</h1>\n'
        for subject_model in grade_model.subjects:
            response += '<h2>' + subject_model.subject_name + ':</h2>\n'
            for student_model in subject_model.students:
                # response += self.grade_trend(student_model)
                student_problem_model = StudentProblemsModel(student_model.fullname)
                response += self.theme_not_understand(student_model, grade_model.subjects)

        return response

    def grade_trend(self, student_model: StudentModel) -> str:
        degree_list = []
        for date_model in student_model.dates:
            if self.check_int(date_model.degree):
                degree_list.append(int(date_model.degree))
        trend = self.linreg(degree_list)
        if trend > 0.5:
            return '   у ' + student_model.fullname + ' наблюдается <strong>повышение</strong> успеваемости\n'
        if trend < 0.5:
            return '   у ' + student_model.fullname + ' наблюдается <strong>понижение</strong> успеваемости\n'
        return ''

    def theme_not_understand(self, student_model: StudentModel, subjects: list) -> str:
        # avg_degree = self.avg_degree(student_model.dates)
        avg_degree_now = []
        response = ''
        problem_models = list()
        for date_model in student_model.dates:
            if self.check_int(date_model.degree):
                avg_degree_now.append(int(date_model.degree))
                if int(round(mean(avg_degree_now), 0)) > int(date_model.degree):
                    # response += f'''\n{date_model.date} {date_model.theme} по причине {date_model.type_of_work}'''
                    problem_model = self.get_problems(student_model, date_model)
                    problem_model = self.analyse_day(problem_model, subjects, student_model.fullname)
                    problem_models.append(problem_model)
        for problem_model in problem_models:
            response += f' {problem_model.date} наблюдается:\n{problem_model.text} \n'
        if response:
            return f'\nУ <strong>{student_model.fullname}</strong> ' + response
        return ''

    def avg_degree(self, dates: list) -> float:
        degree_count = 0
        degree_sum = 0
        for date_model in dates:
            if self.check_int(date_model.degree):
                degree_count += 1
                degree_sum += int(date_model.degree)
        return degree_sum/degree_count

    def check_int(self, s):
        if not s:
            return False
        if s[0] in ('-', '+'):
            return s[1:].isdigit()
        return s.isdigit()

    def linreg(self, x_list) -> float:
        # avg_x = sum(x_list)/len(x_list)
        trends_list = []
        x_last = x_list[0]
        for x in x_list:
            if x > x_last:
                trends_list.append(1)
            if x < x_last:
                trends_list.append(0)
            x_last = x
        if len(trends_list) == 0:
            return 0.5
        return sum(trends_list) / len(trends_list)

    def get_problems(self, student_model: StudentModel, date_model: DateModel) -> ProblemModel:
        date_index = student_model.dates.index(date_model)
        problem_model = ProblemModel(date_model.date, date_model.theme, date_model.type_of_work)
        if date_index > 3:
            if date_model.theme:
                problem_model.text = problem_model.text + f'    Проблема наблюдается на теме {date_model.theme}\n'
            previos_date_model = student_model.dates[date_index-1]
            if previos_date_model.degree == 'Н':
                problem_model.text = problem_model.text + '    Болел и пропустил прошлое занятие\n'
            verbal_count = 0
            verbal_sum = 0
            written_count = 0
            written_sum = 0
            for i in range(date_index):
                temp_date_model = student_model.dates[i]
                if temp_date_model.type_of_work == 'Ответ на уроке' or temp_date_model.type_of_work is None:
                    if self.check_int(temp_date_model.degree):
                        verbal_count += 1
                        verbal_sum += int(temp_date_model.degree)
                else:
                    if self.check_int(temp_date_model.degree):
                        written_count += 1
                        written_sum += int(temp_date_model.degree)
            verba_avg = 0
            written_avg = 0
            if verbal_count:
                verba_avg = verbal_sum / verbal_count
            if written_count:
                written_avg = written_sum/written_count

            if verbal_count and written_count:
                if verba_avg - written_avg > 0.7:
                    problem_model.text = problem_model.text + '    Наблюдается пониженная успеваемость на устных работах\n'
                if written_avg - verba_avg > 0.7:
                    problem_model.text = problem_model.text + '    Наблюдается пониженная успеваемость на письменных работах\n'

        return problem_model

    def analyse_day(self, problem_model: ProblemModel, subjects: list, student_name: str) -> ProblemModel:
        that_day_subjects = []
        for subject in subjects:
            for student in subject.students:
                if student.fullname == student_name:
                    for date in student.dates:
                        if date.date == problem_model.date:
                            that_day_subjects.append(date.type_of_work)
        if 'Контрольная работа' in that_day_subjects:
            problem_model.text = problem_model.text + '    В этот день была Контрольная работа\n'
        if 'Практическая работа' in that_day_subjects:
            problem_model.text = problem_model.text + '    В этот день была Практическая работа\n'
        if 'Самостоятельная работа' in that_day_subjects:
            problem_model.text = problem_model.text + '    В этот день была Самостоятельная работа\n'
        if 'Тестирование' in that_day_subjects:
            problem_model.text = problem_model.text + '    В этот день было Тестирование\n'
        if 'Зачёт' in that_day_subjects:
            problem_model.text = problem_model.text + '    В этот день был Зачёт по физ-ре\n'
        if 'Диктант' in that_day_subjects or 'Сочинение' in that_day_subjects or 'Изложение' in that_day_subjects:
            problem_model.text = problem_model.text + '    В этот день был Диктант/Сочинение/Изложение\n'
        return problem_model



